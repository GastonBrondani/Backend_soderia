from __future__ import annotations

import argparse
import unicodedata
from pathlib import Path
from typing import Dict, Optional

from openpyxl import load_workbook
from fastapi import HTTPException
from sqlalchemy import and_, func, select, update
from sqlalchemy.orm import Session

from app.core.database import SessionLocal
from app.models.cliente import Cliente
from app.models.persona import Persona
from app.models.direccionCliente import DireccionCliente
from app.models.telefonoCliente import TelefonoCliente
from app.models.emailCliente import MailCliente
from app.models.clienteCuenta import ClienteCuenta
from app.models.clienteDiaSemana import ClienteDiaSemana
from app.models.diaSemana import DiaSemana


EMPRESA_ID_DEFAULT = 1
TURNO_DEFAULT = "manana"

# Índices de columnas del Excel v2.0
COL_DNI = 0
COL_NOMBRE = 1
COL_APELLIDO = 2
COL_OBSERVACION = 3
COL_DIRECCION = 4
COL_NRO = 5
COL_TELEFONO = 11
COL_MAIL = 12
COL_DIA_VISITA = 14


DAY_MAP = {
    "lunes": ["lun"],
    "martes": ["mar"],
    "miercoles": ["mie"],
    "miércoles": ["mie"],
    "jueves": ["jue"],
    "viernes": ["vie"],
    "sabado": ["sab"],
    "sábado": ["sab"],
    "domingo": ["dom"],
    "lunes y jueves": ["lun", "jue"],
    "martes y jueves": ["mar", "jue"],
    "lunes y miercoles": ["lun", "mie"],
    "lunes y miércoles": ["lun", "mie"],
    "martes y viernes": ["mar", "vie"],
    "miercoles y viernes": ["mie", "vie"],
    "miércoles y viernes": ["mie", "vie"],
}


def _strip_accents(value: str) -> str:
    return "".join(
        ch
        for ch in unicodedata.normalize("NFKD", value)
        if not unicodedata.combining(ch)
    )


def normalize_text(value: object) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return text


def normalize_key(value: object) -> Optional[str]:
    text = normalize_text(value)
    if text is None:
        return None
    text = _strip_accents(text).lower()
    text = " ".join(text.split())
    return text


def normalize_int_str(value: object) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).strip()
    text = str(value).strip()
    if not text:
        return None
    # Para casos tipo '3434168597.0'
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def normalize_mail(value: object) -> str:
    text = normalize_text(value)
    if text is None:
        return "no tiene"
    text = text.strip()
    if not text:
        return "no tiene"
    return text


def build_address(direccion: object, nro: object) -> Optional[str]:
    calle = normalize_text(direccion)
    numero = normalize_int_str(nro)
    if calle and numero:
        return f"{calle} {numero}".strip()
    return calle or numero


def parse_days(raw_value: object) -> list[str]:
    key = normalize_key(raw_value)
    if key is None:
        raise ValueError("día de visita vacío")
    days = DAY_MAP.get(key)
    if days is None:
        raise ValueError(f"día de visita no reconocido: {raw_value!r}")
    return days


def _idx_dias(db: Session) -> Dict[str, int]:
    dias_db = db.execute(select(DiaSemana)).scalars().all()
    idx: Dict[str, int] = {}
    for d in dias_db:
        nombre = normalize_key(d.nombre_dia)
        if nombre is None:
            continue
        if nombre.startswith("lu"):
            idx["lun"] = d.id_dia
        elif nombre.startswith("ma") and "r" in nombre:
            idx["mar"] = d.id_dia
        elif nombre.startswith("mi"):
            idx["mie"] = d.id_dia
        elif nombre.startswith("ju"):
            idx["jue"] = d.id_dia
        elif nombre.startswith("vi"):
            idx["vie"] = d.id_dia
        elif nombre.startswith("sa"):
            idx["sab"] = d.id_dia
        elif nombre.startswith("do"):
            idx["dom"] = d.id_dia
    return idx


def _calcular_orden_y_correr(
    db: Session,
    id_dia: int,
    turno_val: Optional[str],
    posicion: str,
    despues_de_legajo: Optional[int] = None,
) -> int:
    filtro_base = and_(
        ClienteDiaSemana.id_dia == id_dia,
        (
            ClienteDiaSemana.turno_visita.is_(None)
            if turno_val is None
            else ClienteDiaSemana.turno_visita == turno_val
        ),
    )

    db.execute(select(ClienteDiaSemana.id_cliente).where(filtro_base).with_for_update())

    if posicion == "inicio":
        db.execute(
            update(ClienteDiaSemana)
            .where(filtro_base)
            .values(orden=func.coalesce(ClienteDiaSemana.orden, 0) + 1)
        )
        return 1

    if posicion == "final":
        max_orden = db.execute(
            select(func.coalesce(func.max(ClienteDiaSemana.orden), 0)).where(
                filtro_base
            )
        ).scalar_one()
        return max_orden + 1

    if not despues_de_legajo:
        raise HTTPException(
            status_code=400,
            detail="Falta 'despues_de_legajo' para posicion='despues'.",
        )

    ref_orden = db.execute(
        select(ClienteDiaSemana.orden).where(
            and_(filtro_base, ClienteDiaSemana.id_cliente == despues_de_legajo)
        )
    ).scalar_one_or_none()

    if ref_orden is None:
        raise HTTPException(
            status_code=404,
            detail="Cliente de referencia no existe en ese día/turno.",
        )

    db.execute(
        update(ClienteDiaSemana)
        .where(and_(filtro_base, ClienteDiaSemana.orden >= ref_orden + 1))
        .values(orden=ClienteDiaSemana.orden + 1)
    )
    return ref_orden + 1


def create_cliente_from_row(
    db: Session,
    *,
    dni: int,
    nombre: str,
    apellido: str,
    observacion: Optional[str],
    direccion: str,
    telefono: str,
    mail: str,
    dias_visita: list[str],
    turno_visita: str,
    idx_dias: Dict[str, int],
    empresa_id: int,
) -> int:
    persona = db.get(Persona, dni)
    if not persona:
        persona = Persona(dni=dni, nombre=nombre, apellido=apellido)
        db.add(persona)
        db.flush()

    existente = db.execute(
        select(Cliente).where(Cliente.dni == dni, Cliente.id_empresa == empresa_id)
    ).scalars().first()
    if existente:
        raise ValueError("duplicado_db")

    nuevo = Cliente(id_empresa=empresa_id, observacion=observacion)
    nuevo.persona = persona
    db.add(nuevo)
    db.flush()

    db.add(DireccionCliente(legajo=nuevo.legajo, direccion=direccion))
    db.add(TelefonoCliente(legajo=nuevo.legajo, nro_telefono=telefono))
    db.add(MailCliente(legajo=nuevo.legajo, mail=mail))

    registros: list[ClienteDiaSemana] = []
    for dia in dict.fromkeys(dias_visita):
        id_dia = idx_dias.get(dia)
        if not id_dia:
            raise ValueError(f"día no existe en tabla dia_semana: {dia}")
        orden_nuevo = _calcular_orden_y_correr(
            db=db,
            id_dia=id_dia,
            turno_val=turno_visita,
            posicion="final",
        )
        registros.append(
            ClienteDiaSemana(
                id_cliente=nuevo.legajo,
                id_dia=id_dia,
                turno_visita=turno_visita,
                orden=orden_nuevo,
            )
        )

    if registros:
        db.add_all(registros)

    db.add(ClienteCuenta(legajo=nuevo.legajo))
    db.flush()
    return nuevo.legajo


def import_clientes_excel(
    file_path: str,
    *,
    empresa_id: int = EMPRESA_ID_DEFAULT,
    turno_visita: str = TURNO_DEFAULT,
    apply_changes: bool = False,
) -> None:
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"No existe el archivo: {path}")

    wb = load_workbook(filename=path, data_only=True, read_only=True)
    ws = wb.active

    db = SessionLocal()
    idx_dias = _idx_dias(db)

    total_filas = 0
    creados = 0
    duplicados_archivo = 0
    duplicados_db = 0
    errores = 0
    advertencias_persona = 0
    dnis_vistos: set[int] = set()
    detalles_error: list[str] = []

    try:
        for excel_row_num, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if excel_row_num == 1:
                continue  # encabezado

            dni_str = normalize_int_str(row[COL_DNI] if len(row) > COL_DNI else None)
            if dni_str is None:
                continue

            total_filas += 1

            try:
                dni = int(dni_str)
                if dni in dnis_vistos:
                    duplicados_archivo += 1
                    continue
                dnis_vistos.add(dni)

                nombre = normalize_text(row[COL_NOMBRE] if len(row) > COL_NOMBRE else None)
                apellido = normalize_text(row[COL_APELLIDO] if len(row) > COL_APELLIDO else None)
                observacion = normalize_text(
                    row[COL_OBSERVACION] if len(row) > COL_OBSERVACION else None
                )
                direccion = build_address(
                    row[COL_DIRECCION] if len(row) > COL_DIRECCION else None,
                    row[COL_NRO] if len(row) > COL_NRO else None,
                )
                telefono = normalize_int_str(
                    row[COL_TELEFONO] if len(row) > COL_TELEFONO else None
                )
                mail = normalize_mail(row[COL_MAIL] if len(row) > COL_MAIL else None)
                dias_visita = parse_days(
                    row[COL_DIA_VISITA] if len(row) > COL_DIA_VISITA else None
                )

                if not nombre:
                    raise ValueError("nombre vacío")
                if not apellido:
                    raise ValueError("apellido vacío")
                if not direccion:
                    raise ValueError("dirección vacía")
                if not telefono:
                    raise ValueError("teléfono vacío")

                persona_existente = db.get(Persona, dni)
                if persona_existente and (
                    normalize_key(persona_existente.nombre) != normalize_key(nombre)
                    or normalize_key(persona_existente.apellido) != normalize_key(apellido)
                ):
                    advertencias_persona += 1

                try:
                    create_cliente_from_row(
                        db,
                        dni=dni,
                        nombre=nombre,
                        apellido=apellido,
                        observacion=observacion,
                        direccion=direccion,
                        telefono=telefono,
                        mail=mail,
                        dias_visita=dias_visita,
                        turno_visita=turno_visita,
                        idx_dias=idx_dias,
                        empresa_id=empresa_id,
                    )
                    creados += 1
                except ValueError as inner_exc:
                    if str(inner_exc) == "duplicado_db":
                        duplicados_db += 1
                        db.rollback()
                        continue
                    raise

            except Exception as exc:
                errores += 1
                db.rollback()
                detalles_error.append(f"Fila Excel {excel_row_num}, DNI {dni_str}: {exc}")
                continue

        print("\nRESUMEN IMPORTACIÓN CLIENTES")
        print(f"Archivo: {path}")
        print(f"Empresa ID: {empresa_id}")
        print(f"Turno visita usado: {turno_visita}")
        print(f"Total filas leídas con DNI: {total_filas}")
        print(f"Clientes a crear / creados: {creados}")
        print(f"Duplicados dentro del archivo: {duplicados_archivo}")
        print(f"Duplicados ya existentes en BD: {duplicados_db}")
        print(f"Advertencias por Persona existente con distinto nombre/apellido: {advertencias_persona}")
        print(f"Errores: {errores}")
        print(f"Modo: {'APLICAR' if apply_changes else 'DRY-RUN'}")

        if detalles_error:
            print("\nDETALLE DE ERRORES")
            for item in detalles_error:
                print(f"- {item}")

        if apply_changes:
            db.commit()
            print("\nImportación aplicada correctamente.")
        else:
            db.rollback()
            print("\nDry-run completado. No se guardó ningún cambio.")

    except Exception:
        db.rollback()
        raise
    finally:
        db.close()
        wb.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Importa clientes desde el Excel v2.0 al backend Soderia."
    )
    parser.add_argument(
        "--file",
        required=True,
        help="Ruta al archivo .xlsx dentro del entorno donde corre el script.",
    )
    parser.add_argument(
        "--empresa-id",
        type=int,
        default=EMPRESA_ID_DEFAULT,
        help="ID de empresa a usar. Default: 1",
    )
    parser.add_argument(
        "--turno",
        default=TURNO_DEFAULT,
        choices=["manana", "tarde", "noche"],
        help="Turno de visita para todos los clientes. Default: manana",
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Si se envía, guarda en BD. Si no, corre en dry-run.",
    )
    args = parser.parse_args()

    import_clientes_excel(
        file_path=args.file,
        empresa_id=args.empresa_id,
        turno_visita=args.turno,
        apply_changes=args.apply,
    )
